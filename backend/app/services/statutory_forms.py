"""
Revised statutory forms (Karnataka) — Excel generators.
========================================================
Government-revised formats, added ALONGSIDE the existing reports:

  • FORM-XIV (=FORM-IX)  Attendance Register-cum-Muster Roll
  • FORM-XV  (=FORM-IV)  Register of Wages, OT, Advances, Fines & Deductions
  • FORM-XVI (=FORM-V)   Wage Slip

Each function returns an in-memory .xlsx (BytesIO) with the page set up to fit
(FORM-XV → Legal landscape, FORM-XVI → A4 portrait, FORM-XIV → A3 landscape).
They reuse the same per-employee wage rows the existing Form-B builds, so the
figures always match the rest of the payroll.
"""
import io
import calendar
import datetime as _dt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.pagebreak import Break


F = 'Arial'
_thin = Side(style='thin', color='000000')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
TITLE = Font(name=F, size=13, bold=True)
SUB   = Font(name=F, size=8, italic=True)
H2    = Font(name=F, size=10, bold=True)
HDR   = Font(name=F, size=8, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
GRP_FILL = PatternFill('solid', fgColor='2E75B6')
BODY  = Font(name=F, size=8)
BOLD  = Font(name=F, size=8, bold=True)
INFO  = Font(name=F, size=8.5)
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
Lf = Alignment(horizontal='left', vertical='center', wrap_text=True)
Rt = Alignment(horizontal='right', vertical='center')


def _period(payroll):
    ndays = calendar.monthrange(payroll.year, payroll.month)[1]
    wpf = f"01-{payroll.month:02d}-{payroll.year}"
    wpt = f"{ndays:02d}-{payroll.month:02d}-{payroll.year}"
    return ndays, wpf, wpt


def _est_line(est, with_pan=False):
    parts = [
        f"Name of the Establishment: {est.company_name}"
        + (f" ({est.branch_name})" if getattr(est, 'branch_name', None) else ""),
        f"Name of the Employer: {est.contact_person or est.company_name}",
        f"Name of the Owner: {est.contact_person or est.company_name}",
    ]
    if with_pan:
        parts.append(f"PAN/TAN of the Employer: {getattr(est, 'pan_number', '') or '—'}")
    parts.append(f"Reg. No / LIN: {est.pf_code or est.esic_code or '—'}")
    return "    |    ".join(parts)


def _bytes(wb):
    out = io.BytesIO(); wb.save(out); out.seek(0); return out


def _page(ws, paper, landscape=True):
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.paperSize = paper           # 9=A4, 5=Legal, 8=A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    for a in ('left', 'right', 'top', 'bottom'):
        setattr(ws.page_margins, a, 0.3)
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1


def _box(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


def _merge(ws, r1, c1, r2, c2, val, font, align, fill=None, border=True):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=val)
    cell.font = font; cell.alignment = align
    if fill:
        cell.fill = fill
    if border:
        _box(ws, r1, c1, r2, c2)
    return cell


# ── per-employee "rate of wage" (full sanctioned Basic / DA / Other) ─────────
def _rate_map(entries):
    """entry_id -> (basic, da, other) full sanctioned amounts, from the entry's
    head breakup (falls back to daily rate / gross)."""
    from app.models.payroll import PayrollEntryHead, SalaryHead
    out = {}
    ids = [e.id for e in entries]
    if not ids:
        return out
    rows = (PayrollEntryHead.query
            .filter(PayrollEntryHead.payroll_entry_id.in_(ids))
            .join(SalaryHead, PayrollEntryHead.salary_head_id == SalaryHead.id)
            .add_columns(SalaryHead.short_code, SalaryHead.head_type)
            .all())
    for peh, short_code, head_type in rows:
        if head_type != 'earning':
            continue
        b, d, o = out.get(peh.payroll_entry_id, (0.0, 0.0, 0.0))
        amt = float(peh.full_amount or 0)
        sc = (short_code or '').upper()
        if sc == 'BASIC':
            b += amt
        elif sc == 'DA':
            d += amt
        else:
            o += amt
        out[peh.payroll_entry_id] = (b, d, o)
    return out


# ── shared normalizers (used by BOTH the Excel builders and the print views) ──
def xv_rows(payroll, est, config, entries, rows):
    """Normalized wage-register rows for FORM-XV (Excel + HTML share these)."""
    _, wpf, wpt = _period(payroll)
    ent_by_id = {e.id: e for e in entries}
    ratemap = _rate_map(entries)
    out = []
    tot = {k: 0.0 for k in ('days', 'ot', 'e_basic', 'e_da', 'e_allow', 'e_ot',
                            'e_total', 'epf', 'esic', 'it', 'ins', 'others', 'ded', 'net')}
    for row in rows:
        e = ent_by_id.get(row.get('_entry_id'))
        rb, rda, roth = ratemap.get(row.get('_entry_id'), (0, 0, 0))
        if not rb and not rda and not roth:
            rb = row.get('rate') or 0
        e_basic = row['basic']; e_da = row['da']
        e_allow = row['spl_basic'] + row['hra'] + row['others'] + row['nph']
        e_ot = row['ot_amount']; e_total = row['gross']
        epf = row['pf']; esic = row['esic']; it = row['income_tax']; ins = row['insurance']
        ded = row['total_ded']
        others = max(0, round(ded - epf - esic - it - ins))
        d = dict(sl=row['sl'], code=row['emp_code'], name=row['name'],
                 desig=(e.employee.designation if e else '') or '',
                 dept=(e.employee.department if e else '') or '',
                 duration='Monthly', wperiod=f"{wpf} to {wpt}",
                 days=row['days_worked'], ot=row['ot_days'],
                 rb=round(rb), rda=round(rda), roth=round(roth),
                 e_basic=round(e_basic), e_da=round(e_da), e_allow=round(e_allow),
                 e_ot=round(e_ot), e_total=round(e_total),
                 epf=round(epf), esic=round(esic), it=round(it), ins=round(ins),
                 others=others, ded=round(ded), net=round(row['net_pay']),
                 date_pay=wpt)
        out.append(d)
        for k, v in (('days', row['days_worked']), ('ot', row['ot_days']),
                     ('e_basic', e_basic), ('e_da', e_da), ('e_allow', e_allow),
                     ('e_ot', e_ot), ('e_total', e_total), ('epf', epf), ('esic', esic),
                     ('it', it), ('ins', ins), ('others', others), ('ded', ded), ('net', row['net_pay'])):
            tot[k] += v
    totals = {k: round(v) for k, v in tot.items()}
    return out, totals, wpf, wpt


def xvi_slips(payroll, est, config, entries, rows):
    """Per-employee wage-slip dicts for FORM-XVI."""
    _, wpf, wpt = _period(payroll)
    ent_by_id = {e.id: e for e in entries}
    ratemap = _rate_map(entries)
    slips = []
    for row in rows:
        e = ent_by_id.get(row.get('_entry_id'))
        emp = e.employee if e else None
        rb, rda, roth = ratemap.get(row.get('_entry_id'), (0, 0, 0))
        te = round(row['gross']); td = round(row['total_ded']); net = round(row['net_pay'])
        epf = round(row['pf']); esic = round(row['esic'])
        slips.append(dict(
            name=emp.name if emp else row['name'],
            father=(emp.father_husband_name if emp else '') or '',
            desig=(emp.designation if emp else '') or '',
            uan=(emp.uan_number if emp else '') or '',
            bank=(emp.bank_account_number if emp else '') or '',
            wperiod=f"{wpf} to {wpt}", rb=round(rb), rda=round(rda), roth=round(roth),
            days=row['days_worked'], ot=round(row['ot_amount']),
            gross=te, total_ded=td, epf=epf, esic=esic,
            others=max(0, td - epf - esic), net=net))
    return slips, wpf, wpt


def attendance_rows(payroll, est, config, entries, rows):
    """Per-employee attendance with a day-by-day list for FORM-XIV."""
    ndays, _, _ = _period(payroll)
    ent_by_id = {e.id: e for e in entries}
    rday = 6
    if getattr(config, 'rest_day_type', 'sunday') == 'fixed_day' and getattr(config, 'rest_day_weekday', None) is not None:
        rday = config.rest_day_weekday
    sundays = {d for d in range(1, ndays + 1) if calendar.weekday(payroll.year, payroll.month, d) == rday}
    holidays = set()
    if getattr(payroll, 'holiday_dates', None):
        for d in payroll.holiday_dates.split(','):
            if d.strip().isdigit():
                holidays.add(int(d.strip()))
    out = []
    for i, row in enumerate(rows, 1):
        e = ent_by_id.get(row.get('_entry_id'))
        emp = e.employee if e else None
        present = int(round(row['days_worked'])); worked = 0
        days = []
        for d in range(1, ndays + 1):
            if d in sundays:
                days.append(('W', 'O', 'wo'))
            elif d in holidays:
                days.append(('H', 'H', 'hol'))
            elif worked < present:
                days.append(('09:00', '18:00', 'p')); worked += 1
            else:
                days.append(('A', '', 'ab'))
        out.append(dict(sl=i, code=row['emp_code'], name=emp.name if emp else row['name'],
                        desig=(emp.designation if emp else '') or '',
                        dept=(emp.department if emp else '') or '',
                        days=days, total_days=present, ot=round(row['ot_days'])))
    return out, ndays


# ════════════════════════════════════════════════════════════════════
# FORM-XV — Register of Wages, OT, Advances, Fines and Deductions
# ════════════════════════════════════════════════════════════════════
def build_form_xv(payroll, est, config, entries, rows):
    wb = Workbook(); ws = wb.active; ws.title = "FORM-XV"
    LAST = 34
    ndays, wpf, wpt = _period(payroll)
    ratemap = _rate_map(entries)
    ent_by_id = {e.id: e for e in entries}

    _merge(ws, 1, 1, 1, LAST, "FORM-IV / FORM-XV", TITLE, C, border=False)
    _merge(ws, 2, 1, 2, LAST, "[See clause (ii) of Sub Rule (1) of Rule 51]", SUB, C, border=False)
    _merge(ws, 3, 1, 3, LAST, "REGISTER OF WAGES, OVERTIME, ADVANCES, FINES AND DEDUCTIONS FOR DAMAGE AND LOSS", H2, C, border=False)
    _merge(ws, 4, 1, 4, LAST, _est_line(est, with_pan=True) + f"    |    Wage Period From {wpf}  To  {wpt}", INFO, Lf, border=False)
    ws.row_dimensions[4].height = 34

    single = {1: "Sl. No", 2: "Employee Code Number", 3: "Name of the Employee",
              4: "Designation", 5: "Department",
              6: "Duration of Payment of Wages (Monthly/Fortnightly/Weekly/Piece-rated)",
              7: "Wage Period From  -  To",
              8: "Total Number of Days Worked during the Wage Period",
              9: "Total overtime hours worked / OT production (piece workers)"}
    for c, t in single.items():
        _merge(ws, 6, c, 7, c, t, HDR, C, HDR_FILL)
    _merge(ws, 6, 10, 6, 12, "Rate of Wage", HDR, C, GRP_FILL)
    for c, t in {10: "(a) Basic", 11: "(b) DA", 12: "(c) Other Allowance"}.items():
        _merge(ws, 7, c, 7, c, t, HDR, C, HDR_FILL)
    _merge(ws, 6, 13, 6, 17, "Amount of Wages Earned", HDR, C, GRP_FILL)
    for c, t in {13: "Basic", 14: "DA", 15: "Allowances", 16: "Over Time", 17: "Total Wages Earned"}.items():
        _merge(ws, 7, c, 7, c, t, HDR, C, HDR_FILL)
    _merge(ws, 6, 18, 6, 27, "Deductions", HDR, C, GRP_FILL)
    dedh = {18: "EPF", 19: "ESIC", 20: "Society", 21: "Income Tax", 22: "Insurance",
            23: "Advances", 24: "Recovery of Fine", 25: "Recovery of damages/losses",
            26: "Total Deductions", 27: "Others"}
    for c, t in dedh.items():
        _merge(ws, 7, c, 7, c, t, HDR, C, HDR_FILL)
    tail = {28: "Net Payment", 29: "Date of Payment", 30: "Receipt by employee / Bank Txn ID",
            31: "Nature of acts/omissions for which fine imposed with date",
            32: "Amount of Fine imposed", 33: "Damage/loss caused by neglect/default",
            34: "Signature of Employer / Representative"}
    for c, t in tail.items():
        _merge(ws, 6, c, 7, c, t, HDR, C, HDR_FILL)
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 54

    for c in range(1, LAST + 1):
        cell = ws.cell(row=8, column=c, value=c)
        cell.font = Font(name=F, size=7, italic=True); cell.alignment = C; cell.border = BORDER

    r = 9
    tot = {k: 0.0 for k in ('days', 'ot', 'e_basic', 'e_da', 'e_allow', 'e_ot',
                            'e_total', 'epf', 'esic', 'it', 'ins', 'others', 'ded', 'net')}
    for row in rows:
        e = ent_by_id.get(row.get('_entry_id'))
        rb, rda, roth = ratemap.get(row.get('_entry_id'), (0, 0, 0))
        if not rb and not rda and not roth:
            rb = row.get('rate') or 0
        e_basic = row['basic']; e_da = row['da']
        e_allow = row['spl_basic'] + row['hra'] + row['others'] + row['nph']
        e_ot = row['ot_amount']; e_total = row['gross']
        epf = row['pf']; esic = row['esic']; it = row['income_tax']; ins = row['insurance']
        total_ded = row['total_ded']
        others = max(0, round(total_ded - epf - esic - it - ins))  # incl. PT etc.
        net = row['net_pay']
        vals = [row['sl'], row['emp_code'], row['name'],
                (e.employee.designation if e else '') or '',
                (e.employee.department if e else '') or '',
                "Monthly", f"{wpf} to {wpt}", row['days_worked'], row['ot_days'],
                round(rb), round(rda), round(roth),
                round(e_basic), round(e_da), round(e_allow), round(e_ot), round(e_total),
                round(epf), round(esic), 0, round(it), round(ins), 0, 0, 0,
                round(total_ded), others, round(net),
                wpt, "Bank Transfer", "", 0, "", ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY; cell.border = BORDER
            cell.alignment = Lf if c in (3, 4, 5, 6, 7, 30, 31) else (Rt if c >= 10 else C)
        ws.row_dimensions[r].height = 24
        for k, v in (('days', row['days_worked']), ('ot', row['ot_days']),
                     ('e_basic', e_basic), ('e_da', e_da), ('e_allow', e_allow),
                     ('e_ot', e_ot), ('e_total', e_total), ('epf', epf), ('esic', esic),
                     ('it', it), ('ins', ins), ('others', others), ('ded', total_ded), ('net', net)):
            tot[k] += v
        r += 1

    # TOTAL row
    _merge(ws, r, 1, r, 9, "TOTAL", BOLD, C)
    totmap = {13: 'e_basic', 14: 'e_da', 15: 'e_allow', 16: 'e_ot', 17: 'e_total',
              18: 'epf', 19: 'esic', 21: 'it', 22: 'ins', 26: 'ded', 27: 'others', 28: 'net'}
    for c in range(10, LAST + 1):
        cell = ws.cell(row=r, column=c, value=(round(tot[totmap[c]]) if c in totmap else ''))
        cell.font = BOLD; cell.alignment = Rt; cell.border = BORDER
    ws.row_dimensions[r].height = 22
    r += 1
    _merge(ws, r, 1, r, LAST, "*Note: Required in case register is maintained physically.", SUB, Lf, border=False)

    widths = {1: 5, 2: 11, 3: 22, 4: 12, 5: 12, 6: 15, 7: 17, 8: 9, 9: 10,
              17: 11, 26: 11, 28: 11, 30: 14, 31: 16, 34: 14}
    for c in range(10, 35):
        widths.setdefault(c, 8)
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    _page(ws, 5, landscape=True)             # Legal landscape
    ws.print_area = f"A1:{get_column_letter(LAST)}{r}"
    return _bytes(wb)


# ════════════════════════════════════════════════════════════════════
# FORM-XVI — Wage Slip (one per employee, A4 portrait, 2 per page)
# ════════════════════════════════════════════════════════════════════
def build_form_xvi(payroll, est, config, entries, rows):
    wb = Workbook(); ws = wb.active; ws.title = "FORM-XVI"
    LAST = 6
    ndays, wpf, wpt = _period(payroll)
    ent_by_id = {e.id: e for e in entries}
    ratemap = _rate_map(entries)

    row = 1
    n = len(rows)
    for idx, rw in enumerate(rows):
        e = ent_by_id.get(rw.get('_entry_id'))
        emp = e.employee if e else None
        rb, rda, roth = ratemap.get(rw.get('_entry_id'), (0, 0, 0))
        total_earn = round(rw['gross']); total_ded = round(rw['total_ded']); net = round(rw['net_pay'])
        epf = round(rw['pf']); esic = round(rw['esic'])
        others_ded = max(0, total_ded - epf - esic)
        b = row
        _merge(ws, b, 1, b, LAST, "FORM-V (See Rule 52)  /  FORM-XVI [See Rule 72(2)]", Font(name=F, size=11, bold=True), C, border=False)
        _merge(ws, b + 1, 1, b + 1, LAST, "WAGE SLIP", Font(name=F, size=12, bold=True), C, border=False)
        _merge(ws, b + 2, 1, b + 2, 3, "Date of Issue", BOLD, Lf); _merge(ws, b + 2, 4, b + 2, 6, wpt, BODY, Lf)
        _merge(ws, b + 3, 1, b + 3, 3, "Name of the Establishment", BOLD, Lf); _merge(ws, b + 3, 4, b + 3, 6, est.company_name, BODY, Lf)
        _merge(ws, b + 4, 1, b + 4, 3, "Address", BOLD, Lf); _merge(ws, b + 4, 4, b + 4, 6, est.address or '', BODY, Lf)
        ws.row_dimensions[b + 4].height = 26
        _merge(ws, b + 5, 1, b + 5, 3, "Period", BOLD, Lf); _merge(ws, b + 5, 4, b + 5, 6, f"{calendar.month_name[payroll.month]} {payroll.year}", BODY, Lf)
        _merge(ws, b + 6, 1, b + 6, 1, "S.No", HDR, C, HDR_FILL); _merge(ws, b + 6, 2, b + 6, 3, "Details", HDR, C, HDR_FILL); _merge(ws, b + 6, 4, b + 6, 6, "Particulars", HDR, C, HDR_FILL)
        slip = [
            ("1", "Name of the Employee", emp.name if emp else rw['name']),
            ("2", "Father's / Mother's / Spouse Name", (emp.father_husband_name if emp else '') or ''),
            ("3", "Designation", (emp.designation if emp else '') or ''),
            ("4", "UAN", (emp.uan_number if emp else '') or ''),
            ("5", "Bank Account Number", (emp.bank_account_number if emp else '') or ''),
            ("6", "Wage Period", f"{wpf} to {wpt}"),
            ("7", "Rate of Wages Payable", f"(a) Basic {round(rb):,}   (b) DA {round(rda):,}   (c) Other Allow. {round(roth):,}"),
            ("8", "Total Attendance / Unit of Work Done", f"{rw['days_worked']} days"),
            ("9", "Overtime Wages", f"{round(rw['ot_amount']):,}"),
            ("10", "Gross Wages Payable", f"{total_earn:,}"),
            ("11", "Total Deductions", f"{total_ded:,}   [(a) EPF {epf:,}  (b) ESI {esic:,}  (c) Others {others_ded:,}]"),
            ("12", "Net Wages Paid", f"{net:,}"),
        ]
        rr = b + 7
        for sn, label, val in slip:
            c1 = ws.cell(row=rr, column=1, value=sn); c1.font = BODY; c1.alignment = C; c1.border = BORDER
            _merge(ws, rr, 2, rr, 3, label, BODY, Lf)
            _merge(ws, rr, 4, rr, 6, val, BOLD if sn in ("10", "12") else BODY, Lf)
            rr += 1
        _merge(ws, rr, 1, rr, LAST, f"Amount in Words: Rupees {net:,} only", Font(name=F, size=8, italic=True), Lf)
        rr += 1
        _merge(ws, rr, 1, rr, 3, "Employee Signature", BODY, C); _merge(ws, rr, 4, rr, 6, "Employer / Pay-in-Charge Signature", BODY, C)
        ws.row_dimensions[rr].height = 28
        rr += 1
        _merge(ws, rr, 1, rr, LAST, "*Note: Required in case register is maintained physically.", SUB, Lf, border=False)
        rr += 2
        if (idx + 1) % 2 == 0 and idx != n - 1:
            ws.row_breaks.append(Break(id=rr - 1))
        row = rr

    for c, w in {1: 6, 2: 16, 3: 16, 4: 12, 5: 12, 6: 16}.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    _page(ws, 9, landscape=False)            # A4 portrait
    return _bytes(wb)


# ════════════════════════════════════════════════════════════════════
# FORM-XIV — Attendance Register-cum-Muster Roll (A3 landscape)
# ════════════════════════════════════════════════════════════════════
def build_form_xiv(payroll, est, config, entries, rows):
    wb = Workbook(); ws = wb.active; ws.title = "FORM-XIV"
    ndays, wpf, wpt = _period(payroll)
    day_start = 7
    LAST = 6 + ndays * 2 + 4
    ent_by_id = {e.id: e for e in entries}

    # weekly-off weekday (0=Mon..6=Sun) from config
    rday = 6
    rtype = getattr(config, 'rest_day_type', 'sunday')
    if rtype == 'fixed_day' and getattr(config, 'rest_day_weekday', None) is not None:
        rday = config.rest_day_weekday
    sundays = {d for d in range(1, ndays + 1) if calendar.weekday(payroll.year, payroll.month, d) == rday}
    holidays = set()
    if getattr(payroll, 'holiday_dates', None):
        for d in payroll.holiday_dates.split(','):
            d = d.strip()
            if d.isdigit():
                holidays.add(int(d))
    shift_in, shift_out = "09:00", "18:00"

    _merge(ws, 1, 1, 1, LAST, "FORM-IX  /  FORM-XIV", TITLE, C, border=False)
    _merge(ws, 2, 1, 2, LAST, "(See clause (iii) of Sub Rule (1) of 51)  /  (See Rule 72 (1)(ii))", SUB, C, border=False)
    _merge(ws, 3, 1, 3, LAST, "ATTENDANCE REGISTER-CUM-MUSTER ROLL", H2, C, border=False)
    _merge(ws, 4, 1, 4, LAST, _est_line(est) + f"    |    For the Month of: {calendar.month_name[payroll.month]} {payroll.year}", INFO, Lf, border=False)
    ws.row_dimensions[4].height = 30

    infoh = {1: "S. No", 2: "Employee Code", 3: "Name of the Employee", 4: "Designation",
             5: "Shift", 6: "Place of Work / Section / Dept."}
    for c, t in infoh.items():
        _merge(ws, 6, c, 8, c, t, HDR, C, HDR_FILL)
    _merge(ws, 6, day_start, 6, day_start + ndays * 2 - 1, "Date & Time of Attendance", HDR, C, GRP_FILL)
    for d in range(ndays):
        c = day_start + d * 2
        _merge(ws, 7, c, 7, c + 1, str(d + 1), HDR, C, HDR_FILL)
        for cc, txt in ((c, "IN"), (c + 1, "OUT")):
            cell = ws.cell(row=8, column=cc, value=txt)
            cell.font = Font(name=F, size=6, bold=True, color='FFFFFF'); cell.alignment = C
            cell.fill = HDR_FILL; cell.border = BORDER
    tailc = {LAST - 3: "Total No. of Days Worked", LAST - 2: "Total No. of OT Hours Worked",
             LAST - 1: "Brief details of tour/assignment outside work place, if any",
             LAST: "*Signature of Register Keeper"}
    for c, t in tailc.items():
        _merge(ws, 6, c, 8, c, t, HDR, C, HDR_FILL)
    ws.row_dimensions[6].height = 16; ws.row_dimensions[7].height = 12; ws.row_dimensions[8].height = 24

    r = 9
    for i, rw in enumerate(rows, 1):
        e = ent_by_id.get(rw.get('_entry_id'))
        emp = e.employee if e else None
        vals = [i, rw['emp_code'], emp.name if emp else rw['name'],
                (emp.designation if emp else '') or '', "General",
                (emp.department if emp else '') or '']
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v); cell.font = BODY
            cell.alignment = Lf if c in (3, 4, 6) else C
        worked = 0
        present = int(round(rw['days_worked']))
        for d in range(1, ndays + 1):
            c = day_start + (d - 1) * 2
            if d in sundays:
                mk = ('W', 'O', 'C00000')
            elif d in holidays:
                mk = ('H', 'H', '2E75B6')
            elif worked < present:
                mk = (shift_in, shift_out, None); worked += 1
            else:
                mk = ('A', '', '808080')
            for cc, txt in ((c, mk[0]), (c + 1, mk[1])):
                cell = ws.cell(row=r, column=cc, value=txt)
                cell.font = Font(name=F, size=6, bold=(mk[2] is not None), color=(mk[2] or '000000'))
                cell.alignment = C
        ws.cell(row=r, column=LAST - 3, value=present).font = BOLD
        ws.cell(row=r, column=LAST - 3).alignment = C
        ws.cell(row=r, column=LAST - 2, value=round(rw['ot_days'])).font = BODY
        ws.cell(row=r, column=LAST - 2).alignment = C
        _box(ws, r, 1, r, LAST)
        ws.row_dimensions[r].height = 15
        r += 1
    _merge(ws, r, 1, r, LAST,
           "Legend: W/O = Weekly Off,  H = Holiday,  A = Absent,  09:00/18:00 = standard IN/OUT time.   "
           "*Required in case Register is maintained physically.", SUB, Lf, border=False)

    for c, w in {1: 4, 2: 9, 3: 16, 4: 11, 5: 8, 6: 14}.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for c in range(day_start, day_start + ndays * 2):
        ws.column_dimensions[get_column_letter(c)].width = 3.6
    for c in range(LAST - 3, LAST + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9
    _page(ws, 8, landscape=True)             # A3 landscape
    ws.print_area = f"A1:{get_column_letter(LAST)}{r}"
    return _bytes(wb)
